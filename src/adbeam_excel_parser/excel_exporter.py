from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from adbeam_excel_parser.models import AuditRunSummary, SiteFitStatus

STATUS_HEADER = "AdBeam статус"
REASON_HEADER = "AdBeam причина"
HTTP_HEADER = "AdBeam HTTP"
OUTPUT_SUFFIX = "_audited"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LIGHT_GREEN_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
LIGHT_YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
LIGHT_RED_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")


def export_audit_to_excel(source_file_path: Path, summary: AuditRunSummary, output_file_path: Path | None = None) -> Path:
    if output_file_path is None:
        output_file_path = build_output_path(source_file_path)

    workbook = load_workbook(source_file_path)
    try:
        worksheet = workbook[summary.sheet_name]
        worksheet.insert_cols(1, amount=3)

        worksheet.cell(row=1, column=1, value=STATUS_HEADER)
        worksheet.cell(row=1, column=2, value=REASON_HEADER)
        worksheet.cell(row=1, column=3, value=HTTP_HEADER)

        for column_index in (1, 2, 3):
            cell = worksheet.cell(row=1, column=column_index)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        worksheet.freeze_panes = "D2"
        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 48
        worksheet.column_dimensions["C"].width = 14

        results_by_row = {result.row_index: result for result in summary.results}
        max_column = worksheet.max_column

        for row_index in range(2, worksheet.max_row + 1):
            result = results_by_row.get(row_index)
            if result is None:
                continue

            worksheet.cell(row=row_index, column=1, value=result.status.value)
            worksheet.cell(row=row_index, column=2, value=result.status_reason)
            worksheet.cell(row=row_index, column=3, value=result.http_status)

            fill = pick_row_fill(result.status)
            for column_index in range(1, max_column + 1):
                worksheet.cell(row=row_index, column=column_index).fill = fill

        workbook.save(output_file_path)
        return output_file_path
    finally:
        workbook.close()


def build_output_path(source_file_path: Path) -> Path:
    return source_file_path.with_name(f"{source_file_path.stem}{OUTPUT_SUFFIX}{source_file_path.suffix}")


def pick_row_fill(status: SiteFitStatus) -> PatternFill:
    if status == SiteFitStatus.FIT_NOW:
        return LIGHT_GREEN_FILL

    if status == SiteFitStatus.FIT_LATER:
        return LIGHT_YELLOW_FILL

    return LIGHT_RED_FILL
