from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from adbeam_excel_parser.models import ExcelReadSummary, ExcelWebsiteRows, RowPreview, WebsiteRow

WEBSITE_HEADER_KEYWORDS = (
    "site",
    "website",
    "web",
    "url",
    "домен",
    "сайт",
    "ссылка",
)

URL_PREFIXES = ("http://", "https://", "www.")


def read_excel_summary(file_path: Path) -> ExcelReadSummary:
    workbook = _open_workbook(file_path)

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)

        if header_row is None:
            raise ValueError("Excel file is empty")

        headers = [normalize_header(value, index) for index, value in enumerate(header_row, start=1)]
        website_column_indexes = detect_website_columns(headers)

        total_rows = 0
        rows_with_websites = 0
        preview: list[RowPreview] = []

        for excel_row_index, row_values in enumerate(rows_iter, start=2):
            total_rows += 1
            row_dict = build_row_dict(headers, row_values)
            website_value = find_website_value(row_values, website_column_indexes)

            if website_value:
                rows_with_websites += 1

            if len(preview) < 5:
                preview.append(
                    RowPreview(
                        row_index=excel_row_index,
                        website=website_value,
                        values=row_dict,
                    )
                )

        return ExcelReadSummary(
            file_path=str(file_path),
            sheet_name=worksheet.title,
            total_rows=total_rows,
            headers=headers,
            website_columns=[headers[index] for index in website_column_indexes],
            rows_with_websites=rows_with_websites,
            preview=preview,
        )
    finally:
        workbook.close()


def extract_website_rows(file_path: Path) -> ExcelWebsiteRows:
    workbook = _open_workbook(file_path)

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)

        if header_row is None:
            raise ValueError("Excel file is empty")

        headers = [normalize_header(value, index) for index, value in enumerate(header_row, start=1)]
        website_column_indexes = detect_website_columns(headers)

        total_rows = 0
        rows: list[WebsiteRow] = []

        for excel_row_index, row_values in enumerate(rows_iter, start=2):
            total_rows += 1
            row_dict = build_row_dict(headers, row_values)
            website_value = find_website_value(row_values, website_column_indexes)

            if not website_value:
                continue

            rows.append(
                WebsiteRow(
                    row_index=excel_row_index,
                    website=website_value,
                    values=row_dict,
                )
            )

        return ExcelWebsiteRows(
            file_path=str(file_path),
            sheet_name=worksheet.title,
            headers=headers,
            total_rows=total_rows,
            website_columns=[headers[index] for index in website_column_indexes],
            rows=rows,
        )
    finally:
        workbook.close()


def _open_workbook(file_path: Path):
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if file_path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported at this step")

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        return load_workbook(filename=file_path, read_only=True, data_only=True)


def normalize_header(value: Any, column_index: int) -> str:
    text = normalize_cell(value)
    return text or f"column_{column_index}"


def build_row_dict(headers: list[str], row_values: tuple[Any, ...]) -> dict[str, Any]:
    result: dict[str, Any] = {}

    for header, raw_value in zip(headers, row_values):
        if raw_value is None:
            result[header] = None
        elif isinstance(raw_value, str):
            result[header] = raw_value.strip()
        else:
            result[header] = raw_value

    return result


def detect_website_columns(headers: list[str]) -> list[int]:
    result: list[int] = []

    for index, header in enumerate(headers):
        normalized = header.strip().lower()
        if any(keyword in normalized for keyword in WEBSITE_HEADER_KEYWORDS):
            result.append(index)

    return result


def find_website_value(row_values: tuple[Any, ...], website_column_indexes: list[int]) -> str | None:
    for column_index in website_column_indexes:
        if column_index >= len(row_values):
            continue

        value = normalize_cell(row_values[column_index])
        if looks_like_website(value):
            return value

    for raw_value in row_values:
        value = normalize_cell(raw_value)
        if looks_like_website(value):
            return value

    return None


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def looks_like_website(value: str) -> bool:
    if not value:
        return False

    lower_value = value.lower()

    if lower_value.startswith(URL_PREFIXES):
        return True

    return "." in lower_value and " " not in lower_value and len(lower_value) >= 4
